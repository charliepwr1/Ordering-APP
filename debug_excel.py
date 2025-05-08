#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl.utils.cell import get_column_letter
import os
import sys

def analyze_excel_file(file_path):
    print(f"Analyzing Excel file: {file_path}")
    print(f"File exists: {os.path.exists(file_path)}")
    print(f"File size: {os.path.getsize(file_path)} bytes")
    print("\n" + "="*80 + "\n")

    # Get sheet names using openpyxl
    print("SHEET NAMES (using openpyxl):")
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        print(f"Found {len(sheet_names)} sheets: {sheet_names}")
        
        # Analyze each sheet
        for sheet_name in sheet_names:
            print("\n" + "="*80)
            print(f"ANALYZING SHEET: '{sheet_name}'")
            print("="*80)
            
            sheet = workbook[sheet_name]
            
            # Get dimensions
            min_row, min_col, max_row, max_col = 1, 1, sheet.max_row, sheet.max_column
            print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
            
            # Look for header rows with AGLC SKU or EachesPerCase
            print("\nSEARCHING FOR HEADER ROWS:")
            header_rows = []
            for row_idx in range(1, min(max_row, 30)):  # Check first 30 rows
                row_values = [cell.value for cell in sheet[row_idx] if cell.value is not None]
                row_str = " ".join([str(val).lower() for val in row_values if val])
                if any(keyword.lower() in row_str for keyword in ["aglc", "sku", "each", "case", "eachespercase"]):
                    header_rows.append(row_idx)
                    print(f"Potential header at row {row_idx}: {row_values}")
            
            # Show detailed cell analysis for the first 15 rows
            print("\nDETAILED CELL ANALYSIS (first 15 rows):")
            for row_idx in range(1, min(max_row, 15) + 1):
                row_info = []
                for col_idx in range(1, min(max_col, 20) + 1):  # Limit to 20 columns
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_addr = f"{get_column_letter(col_idx)}{row_idx}"
                    cell_value = cell.value
                    cell_type = type(cell_value).__name__ if cell_value is not None else "None"
                    row_info.append(f"{cell_addr}:{cell_value}({cell_type})")
                print(f"Row {row_idx}: {' | '.join(row_info)}")
    
    except Exception as e:
        print(f"Error in openpyxl analysis: {str(e)}")
    
    # Try using pandas
    print("\n" + "="*80)
    print("PANDAS ANALYSIS:")
    print("="*80)
    try:
        # Get sheet names using pandas
        excel_file = pd.ExcelFile(file_path)
        sheet_names_pd = excel_file.sheet_names
        print(f"Pandas found {len(sheet_names_pd)} sheets: {sheet_names_pd}")
        
        for sheet_name in sheet_names_pd:
            print(f"\nANALYZING SHEET WITH PANDAS: '{sheet_name}'")
            
            # Read sheet with no header first to see raw data
            df_no_header = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
            print(f"\nRaw data (first 10 rows, no header):")
            print(df_no_header)
            
            # Try with header=0
            df_header0 = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10)
            print(f"\nWith header=0:")
            print(df_header0)
            print("\nColumn names:", df_header0.columns.tolist())
            
            # Check data types
            print("\nData types:")
            print(df_header0.dtypes)
            
            # Try to find AGLC SKU and EachesPerCase in raw data
            print("\nSearching for 'AGLC', 'SKU', and 'Eaches' in data:")
            search_terms = ["aglc", "sku", "each", "case", "eachespercase"]
            
            # Convert all values to string for searching
            df_str = df_no_header.astype(str).apply(lambda x: x.str.lower())
            
            for term in search_terms:
                matches = (df_str == term) | df_str.apply(lambda x: x.str.contains(term, na=False))
                if matches.any().any():
                    print(f"Found '{term}' at following positions:")
                    for i, row in enumerate(matches.values):
                        for j, val in enumerate(row):
                            if val:
                                print(f"  Row {i}, Column {j}")
    
    except Exception as e:
        print(f"Error in pandas analysis: {str(e)}")

if __name__ == "__main__":
    # Use provided file path or default
    file_path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/c/Users/charl/Projects/Cannabis-order-app/CannabisRetailersManualOrderForm.xlsm"
    analyze_excel_file(file_path)