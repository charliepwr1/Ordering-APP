#!/usr/bin/env python3
"""
Utility to check if a file is a valid Excel document and contains the expected sheets.
"""
import os
import sys
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

def check_excel_file(file_path):
    """
    Verify that a file is a valid Excel document and has expected structure.
    Returns True if valid, False otherwise.
    """
    try:
        # Check file exists
        if not os.path.exists(file_path):
            print(f"Error: File {file_path} does not exist.")
            return False
            
        # Check file size
        file_size = os.path.getsize(file_path)
        print(f"File size: {file_size} bytes")
        if file_size < 10000:  # Excel files are typically larger than 10KB
            print("Warning: File is unusually small for an Excel document.")
        
        # Check file signature
        with open(file_path, 'rb') as f:
            header = f.read(8)
            if not header.startswith(b"PK"):
                print("Error: File doesn't have Excel/ZIP file signature (PK).")
                print(f"First 8 bytes: {header}")
                return False
        
        # Try opening with openpyxl
        try:
            wb = load_workbook(filename=file_path, read_only=True)
            print(f"Sheets in workbook: {wb.sheetnames}")
            
            # Check for either "Catalog" or "Catalogue" sheet
            if 'Catalog' not in wb.sheetnames and 'Catalogue' not in wb.sheetnames:
                print("Warning: Neither 'Catalog' nor 'Catalogue' sheet found.")
                print(f"Available sheets: {wb.sheetnames}")
            else:
                # Found at least one of the sheets - prioritize "Catalogue"
                sheet_name = 'Catalogue' if 'Catalogue' in wb.sheetnames else 'Catalog'
                print(f"Found expected sheet: '{sheet_name}'")
            return True
        except Exception as e:
            print(f"Error opening with openpyxl: {str(e)}")
        
        # Try opening with pandas as fallback
        try:
            xls = pd.ExcelFile(file_path)
            print(f"Sheets in workbook (pandas): {xls.sheet_names}")
            
            # Check for either "Catalog" or "Catalogue" sheet
            if 'Catalog' not in xls.sheet_names and 'Catalogue' not in xls.sheet_names:
                print("Warning: Neither 'Catalog' nor 'Catalogue' sheet found in pandas.")
                print(f"Available sheets (pandas): {xls.sheet_names}")
            else:
                # Found at least one of the sheets - prioritize "Catalogue"
                sheet_name = 'Catalogue' if 'Catalogue' in xls.sheet_names else 'Catalog'
                print(f"Found expected sheet (pandas): '{sheet_name}'")
            return True
        except Exception as e:
            print(f"Error opening with pandas: {str(e)}")
            
        return False
    except Exception as e:
        print(f"Unexpected error checking file: {str(e)}")
        return False

if __name__ == "__main__":
    # Check command line arguments
    if len(sys.argv) != 2:
        file_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "CannabisRetailersManualOrderForm.xlsm"
        )
        print(f"No file specified, checking default: {file_path}")
    else:
        file_path = sys.argv[1]
    
    # Check the file
    print(f"Checking file: {file_path}")
    if check_excel_file(file_path):
        print("✅ File appears to be a valid Excel document.")
    else:
        print("❌ File does not appear to be a valid Excel document.")