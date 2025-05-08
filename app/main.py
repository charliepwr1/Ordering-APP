import os
import sys
import numpy as np

#  ────────────────────────────────────────────────────────────
#  Make sure the parent folder (project root) is on sys.path
#  so that "import etl.generate_order" works.
#  
#  __file__ is ".../cannabis-order-app/app/main.py", so:
#    parent = .../cannabis-order-app
#  ────────────────────────────────────────────────────────────
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if project_root not in sys.path:
    sys.path.insert(0, project_root)


import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from datetime import datetime
from etl.generate_order import generate_order
from download_order_form import download_order_form  # your helper

st.title("Cannabis Order Generator")

st.markdown("""
## Alberta Cannabis Order Generator

This app downloads the AGLC order form template and automatically populates it with current inventory data.
For each location, a separate sheet will be created with the complete order form.
""")

col1, col2, col3 = st.columns(3)
with col1:
    hist_days = st.number_input("Days of historical IOH", min_value=1, max_value=90, value=30)
with col2:
    exclude_today = st.checkbox("Exclude today from selected range")
with col3:
    receiving_date = st.date_input("Expected Receiving Date", value=pd.to_datetime('today') + pd.Timedelta(days=7))

st.markdown("""
**Order Calculation Parameters:**
- Orders will be calculated to cover inventory needs from now until **14 days after** the receiving date
- Formula: `Sales/day × (days until receiving + 14) - (In Stock + On Order)`
- This will be divided by case size to determine cases needed
""")

st.divider()

if st.button("Run ETL & Prepare Compiled Order Form"):
    with st.spinner("Running ETL process..."):
        # 1) Run your ETL, which writes 'output/Final_Report.xlsx'
        output_path = "output/Final_Report.xlsx"
        generate_order(output_path, hist_days=hist_days, exclude_today=exclude_today)
        st.success("✅ ETL complete – got inventory & sales data.")

    # 2) Try to download the blank order-form into memory
    try:
        order_form_bytes = download_order_form()
        st.info("🔄 Downloaded blank order-form template")
    except Exception as e:
        st.warning(f"⚠️ Automatic download failed: {str(e)}")
        
        # Check if we have a local copy of the file
        local_path = os.path.join(project_root, "CannabisRetailersManualOrderForm.xlsm")
        if os.path.exists(local_path):
            # Verify file is valid Excel
            try:
                from app.check_excel import check_excel_file
                is_valid = check_excel_file(local_path)
                if is_valid:
                    st.info("Using locally stored template file instead")
                    with open(local_path, "rb") as f:
                        order_form_bytes = f.read()
                else:
                    st.warning(f"Local file at {local_path} doesn't appear to be a valid Excel file. It may be HTML instead.")
                    raise ValueError("Invalid Excel file")
            except Exception as e:
                st.warning(f"Error validating local file: {str(e)}")
                raise
        else:
            st.error(
                "No local template file found or the file is not a valid Excel document. "
                "The website now requires CAPTCHA verification. "
                "Please download the form manually and place it in the project root folder."
            )
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("Simple Download Guide"):
                    from simple_download_guide import display_manual_download_instructions
                    display_manual_download_instructions()
                    st.info("Follow the printed instructions in your terminal/console window.")
                    st.info("After downloading, restart this app.")
            
            with col2:
                if st.button("Detailed Website Guide"):
                    from website_guide import display_website_guide
                    display_website_guide()
                    st.info("Follow the detailed guide in your terminal/console window.")
                    st.info("This will help diagnose website changes.")
            
            # Add information about the exact file location
            st.info(f"Place the downloaded file at: {os.path.abspath(os.path.join(project_root, 'CannabisRetailersManualOrderForm.xlsm'))}")
            
            # Add a diagnostic button to check sheet names in a manually uploaded file
            if st.button("Check Local File Sheets"):
                check_path = os.path.join(project_root, "CannabisRetailersManualOrderForm.xlsm")
                if os.path.exists(check_path):
                    try:
                        from app.check_excel import check_excel_file
                        if check_excel_file(check_path):
                            # Load workbook to examine sheets
                            try:
                                wb = load_workbook(filename=check_path, read_only=True)
                                st.success(f"Found sheets: {wb.sheetnames}")
                                for sheet in wb.sheetnames:
                                    st.code(f"Sheet: {sheet}")
                            except Exception as e:
                                st.error(f"Error examining sheets: {str(e)}")
                        else:
                            st.error("The local file doesn't appear to be a valid Excel file.")
                    except Exception as e:
                        st.error(f"Error checking file: {str(e)}")
                else:
                    st.warning(f"No file found at {check_path}")
            
            # Option to try to use an existing file
            upload_file = st.file_uploader("Or upload the Excel file directly:", type=["xlsm"])
            if upload_file is not None:
                # Get the uploaded file data
                bytes_data = upload_file.getvalue()
                
                # Do a more thorough check to ensure it's a valid Excel file
                if bytes_data.startswith(b"PK"):
                    # Save the file to the project root
                    file_path = os.path.join(project_root, "CannabisRetailersManualOrderForm.xlsm")
                    with open(file_path, "wb") as f:
                        f.write(bytes_data)
                    
                    # Verify using the check_excel utility
                    from app.check_excel import check_excel_file
                    is_valid = check_excel_file(file_path)
                    
                    if is_valid:
                        st.success(f"✅ Valid Excel file uploaded and saved to {file_path}")
                        # Set order_form_bytes and continue processing
                        order_form_bytes = bytes_data
                        # Fall through to continue processing
                    else:
                        st.error("❌ The file has the ZIP signature but doesn't appear to be a valid Excel file with the expected structure.")
                        st.warning("Please upload a proper Cannabis Retailers Manual Order Form (.xlsm file).")
                        st.stop()
                else:
                    st.error("❌ The uploaded file doesn't appear to be a valid Excel file.")
                    st.stop()
            else:
                st.stop()

    # 3) Load the “Catalogue” sheet (cols A–F from row 11) safely
    buf = io.BytesIO(order_form_bytes)
    buf.seek(0)
    # sanity check: should start with PK for a ZIP-based Office file
    if not order_form_bytes.startswith(b"PK"):
        st.error(
            "Download didn’t return a valid Excel file. "
            "First 200 bytes:\n\n"
            + order_form_bytes[:200].decode("utf-8", errors="replace")
        )
        st.stop()

    try:
        # Try the different possible sheet names - prioritize "Catalogue" as that's the correct spelling
        sheet_names_to_try = ["Catalogue", "Catalog"]
        catalogue_df = None
        last_error = None
        
        for sheet_name in sheet_names_to_try:
            try:
                # The issue appears to be that the headers are not in the first row
                # Try different header row positions to find the actual data
                for header_row in range(0, 20):  # Try first 20 rows as possible headers
                    try:
                        temp_df = pd.read_excel(
                            buf,
                            sheet_name=sheet_name,
                            engine="openpyxl",
                            header=header_row,
                        )
                        
                        # Check if this looks like a real header row (no Unnamed columns)
                        unnamed_count = sum(1 for col in temp_df.columns if 'Unnamed' in str(col))
                        if unnamed_count < len(temp_df.columns) / 2:  # Less than half are unnamed
                            # This might be a good header row
                            st.info(f"Trying header row {header_row}: {temp_df.columns.tolist()}")
                            
                            # Look for our important columns
                            if any('AGLC SKU' == str(col).strip() for col in temp_df.columns):
                                st.success(f"✅ Found AGLC SKU in header row {header_row}")
                                catalogue_df = temp_df
                                break
                                
                            if any('EachesPerCase' == str(col).strip() for col in temp_df.columns):
                                st.success(f"✅ Found EachesPerCase in header row {header_row}")
                                catalogue_df = temp_df
                                break
                                
                            # Also check case-insensitive
                            if any('aglc sku' in str(col).lower().strip() for col in temp_df.columns) or \
                               any('eachespercase' in str(col).lower().strip() for col in temp_df.columns):
                                st.success(f"✅ Found key columns (case insensitive) in header row {header_row}")
                                catalogue_df = temp_df
                                break
                    except Exception as row_err:
                        # This row didn't work as a header, continue to the next
                        continue
                
                # If we still don't have a dataframe, try the default first row
                if 'catalogue_df' not in locals() or catalogue_df is None:
                    catalogue_df = pd.read_excel(
                        buf,
                        sheet_name=sheet_name,
                        engine="openpyxl",
                    )
                
                # Debug output to verify we're getting EachesPerCase
                st.info(f"All columns found in order form: {catalogue_df.columns.tolist()}")
                if 'EachesPerCase' in catalogue_df.columns:
                    st.success(f"✅ EachesPerCase column FOUND in sheet {sheet_name}")
                else:
                    st.warning(f"⚠️ EachesPerCase NOT found in columns. Available columns: {catalogue_df.columns.tolist()}")
                st.info(f"Successfully loaded sheet: '{sheet_name}'")
                break  # Break the loop if successful
            except Exception as e:
                last_error = e
                buf.seek(0)  # Reset buffer position for next attempt
        
        if catalogue_df is None:
            raise last_error or ValueError("Could not find either 'Catalog' or 'Catalogue' sheet")
            
    except Exception as e:
        # Fallback: manually load via openpyxl with different approach
        st.warning(f"Pandas loading failed: {str(e)}. Trying manual openpyxl loading...")
        buf.seek(0)
        wb = load_workbook(filename=buf, keep_vba=False, read_only=True)
        
        # Try different sheet names, prioritizing "Catalogue"
        sheet_name = None
        for name in ["Catalogue", "Catalog"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
                
        if not sheet_name:
            st.error(f"Could not find 'Catalog' or 'Catalogue' sheet. Available sheets: {wb.sheetnames}")
            raise ValueError(f"Required sheet not found. Available: {wb.sheetnames}")
            
        ws = wb[sheet_name]
        
        # Get all row data and find the header row by looking for "EachesPerCase"
        header_row_index = None
        header_row = None
        
        # Show the first row to help debugging
        first_row = [cell.value for cell in list(ws.rows)[0]]
        st.info(f"First row of the sheet: {first_row}")
        
        # Look through the first 20 rows to find headers
        for row_idx in range(1, 21):
            try:
                # Get the cells in this row
                row_data = [cell.value for cell in list(ws.rows)[row_idx-1]]
                
                # See if this looks like a header row
                if row_data and any(isinstance(cell, str) and cell.strip() == "EachesPerCase" for cell in row_data):
                    header_row_index = row_idx
                    header_row = row_data
                    st.success(f"✅ Found header row with exact match for EachesPerCase at row {row_idx}")
                    break
                    
                # Also check case-insensitive
                row_data_lower = [str(cell).lower() if cell is not None else "" for cell in row_data]
                if any("eachespercase" == str(cell).lower().strip() for cell in row_data):
                    header_row_index = row_idx
                    header_row = row_data
                    pos = next(i for i, cell in enumerate(row_data) if str(cell).lower().strip() == "eachespercase")
                    st.success(f"✅ Found header row with 'eachespercase' (case-insensitive) at row {row_idx}, position {pos}")
                    break
            except Exception as row_error:
                st.warning(f"Error processing row {row_idx}: {str(row_error)}")
                continue
                
        if header_row_index:
            try:
                # Get data rows (all rows after the header)
                data_rows = []
                for row_idx, row in enumerate(list(ws.rows)[header_row_index:]):
                    try:
                        row_data = [cell.value for cell in row]
                        if any(cell is not None for cell in row_data):  # Skip empty rows
                            data_rows.append(row_data)
                    except Exception as row_err:
                        st.warning(f"Error in data row {row_idx}: {str(row_err)}")
                        continue
                        
                # Create DataFrame with all columns from the header
                catalogue_df = pd.DataFrame(data_rows, columns=header_row)
                st.info(f"Loaded sheet '{sheet_name}' manually via openpyxl with all columns including EachesPerCase")
                
                # Verify EachesPerCase is there
                eaches_col = next((col for col in catalogue_df.columns if str(col).lower().strip() == "eachespercase"), None)
                if eaches_col:
                    st.success(f"✅ EachesPerCase found as '{eaches_col}' in manual loading approach")
                    # Show a sample
                    st.info(f"Sample values: {catalogue_df[eaches_col].head().tolist()}")
                    # Rename to standard form if needed
                    if eaches_col != "EachesPerCase":
                        catalogue_df["EachesPerCase"] = catalogue_df[eaches_col]
                        st.info(f"Created standardized EachesPerCase column from '{eaches_col}'")
                else:
                    st.warning(f"⚠️ EachesPerCase still not found after manual loading. Available columns: {catalogue_df.columns.tolist()}")
            except Exception as df_error:
                st.error(f"Error creating DataFrame from manual load: {str(df_error)}")
                # Fall through to default approach
                header_row_index = None
                
        if not header_row_index:
            # Fallback to original approach if we can't find header row
            st.warning("Could not find header row with EachesPerCase, using standard row positions")
            try:
                # Get ALL columns from the file, not just A-F
                all_rows = list(ws.rows)
                if len(all_rows) > 0:
                    num_cols = len(all_rows[0])
                    st.info(f"Found {num_cols} columns in the sheet")
                    
                    # Try to read all columns from the first row
                    headers = [cell.value for cell in all_rows[0]]
                    if 'EachesPerCase' in headers:
                        st.success(f"✅ Found EachesPerCase in first row at position {headers.index('EachesPerCase')}")
                    else:
                        st.warning(f"First row doesn't contain EachesPerCase: {headers}")
                        
                    # Get all data from subsequent rows
                    data = []
                    for row in all_rows[1:]:
                        if any(cell.value is not None for cell in row):
                            data.append([cell.value for cell in row])
                    
                    catalogue_df = pd.DataFrame(data, columns=headers)
                else:
                    st.error("No rows found in the sheet")
                    # Create an empty DataFrame as fallback
                    catalogue_df = pd.DataFrame()
            except Exception as fallback_error:
                st.error(f"Error in fallback approach: {str(fallback_error)}")
                # Create a minimal DataFrame to avoid total failure
                catalogue_df = pd.DataFrame()

    # 4) Load data from the ETL output (which creates sheets by location, not named "Weekly")
    # First, check what sheets are available in the output file
    available_sheets = pd.ExcelFile(output_path).sheet_names
    st.info(f"Available sheets in ETL output: {available_sheets}")
    
    if not available_sheets:
        st.error("No sheets found in the ETL output file. Check the generate_order.py script.")
        st.stop()
    
    # Load all sheets except 'Summary' and combine them
    relevant_sheets = [sheet for sheet in available_sheets if sheet.lower() != 'summary']
    
    if not relevant_sheets:
        st.error("No usable data sheets found in the ETL output file.")
        st.stop()
    
    # Load all sheets
    all_dfs = {}
    for sheet in relevant_sheets:
        df = pd.read_excel(output_path, sheet_name=sheet)
        st.info(f"Sheet '{sheet}' contains {len(df)} rows with columns: {df.columns.tolist()}")
        all_dfs[sheet] = df
    
    # Combine all data frames, but first check if they have compatible columns
    common_columns = set.intersection(*[set(df.columns) for df in all_dfs.values()]) if all_dfs else set()
    st.info(f"Common columns across all sheets: {sorted(list(common_columns))}")
    
    # Combine the data frames into one
    sheet_dfs = list(all_dfs.values())
    if len(sheet_dfs) > 1:
        weekly_df = pd.concat(sheet_dfs, ignore_index=True)
        st.success(f"Combined {len(sheet_dfs)} sheets into one dataset with {len(weekly_df)} rows")
    else:
        weekly_df = sheet_dfs[0]
        st.info(f"Using single sheet '{relevant_sheets[0]}' with {len(weekly_df)} rows")

    # 5) Merge on AGLC SKU and Supplier SKU
    # Show available columns in both dataframes for debugging
    st.info(f"Catalogue columns: {catalogue_df.columns.tolist()}")
    st.info(f"ETL output columns: {weekly_df.columns.tolist()}")
    
    # Look for the expected column names
    catalogue_sku_col = None
    weekly_sku_col = None
    stock_qty_col = None
    
    # CRITICAL: If the catalogue_df still has mostly unnamed columns or no rows, 
    # we need to create a proper structure for it
    if (sum(1 for col in catalogue_df.columns if 'Unnamed' in str(col)) > len(catalogue_df.columns) / 2) or len(catalogue_df) == 0:
        st.warning("Order form appears to have an unexpected structure. Creating a synthetic catalogue dataframe.")
        
        # Create a basic dataframe with the structure we need
        synthetic_catalogue = {
            'AGLC SKU': weekly_df['SKU'].tolist(),  # Use SKUs from the ETL data
            'Brand Name': weekly_df['Brand'].tolist() if 'Brand' in weekly_df.columns else [''] * len(weekly_df),
            'Product': weekly_df['Product'].tolist(),
            'EachesPerCase': [12] * len(weekly_df),  # Default to 12 units per case
            'Format': weekly_df['Classification'].tolist() if 'Classification' in weekly_df.columns else [''] * len(weekly_df),
        }
        
        # Replace the problematic catalogue_df with our synthetic one
        catalogue_df = pd.DataFrame(synthetic_catalogue)
        st.success("Created synthetic catalogue dataframe with required columns")
    
    # Check for AGLC SKU in catalogue_df
    aglc_sku_alternatives = [
        "AGLC SKU", "SKU", "Product Code", "Item Code", "AGLC Code", "Product SKU"
    ]
    for col in aglc_sku_alternatives:
        if col in catalogue_df.columns:
            catalogue_sku_col = col
            break
    
    # Check for Supplier SKU in weekly_df
    supplier_sku_alternatives = [
        "Supplier SKU", "CNFR SKU", "CNB SKU", "CNB-SKU", "Supplier Code"
    ]
    for col in supplier_sku_alternatives:
        if col in weekly_df.columns:
            weekly_sku_col = col
            break
    
    # If still not found, try any column with "SKU" in the name
    if not catalogue_sku_col:
        sku_cols = [col for col in catalogue_df.columns if "sku" in col.lower()]
        if sku_cols:
            catalogue_sku_col = sku_cols[0]
            
    if not weekly_sku_col:
        sku_cols = [col for col in weekly_df.columns if "sku" in col.lower()]
        if sku_cols:
            weekly_sku_col = sku_cols[0]
    
    # Look for In Stock Qty column
    stock_alternatives = [
        "In Stock Qty", "Stock Qty", "Current Stock", "Quantity", "Qty", "In Stock"
    ]
    for col in stock_alternatives:
        if col in weekly_df.columns:
            stock_qty_col = col
            break
    
    # If still not found, try any column with "stock" or "qty" in the name
    if not stock_qty_col:
        stock_cols = [col for col in weekly_df.columns if "stock" in col.lower() or "qty" in col.lower()]
        if stock_cols:
            stock_qty_col = stock_cols[0]
    
    # Show what columns we're using
    if catalogue_sku_col and weekly_sku_col and stock_qty_col:
        st.success(f"Merging on: Catalogue['{catalogue_sku_col}'] = ETL['{weekly_sku_col}'], using '{stock_qty_col}' for stock")
    else:
        st.error("Could not identify required columns for merging")
        if not catalogue_sku_col:
            st.error(f"Missing SKU column in Catalogue. Available: {catalogue_df.columns.tolist()}")
        if not weekly_sku_col:
            st.error(f"Missing Supplier SKU column in ETL output. Available: {weekly_df.columns.tolist()}")
        if not stock_qty_col:
            st.error(f"Missing In Stock Qty column in ETL output. Available: {weekly_df.columns.tolist()}")
        
        # Try to continue with defaults if needed
        if not catalogue_sku_col:
            # If we have Unnamed columns and no proper SKU column, create one
            if 'AGLC SKU' not in catalogue_df.columns:
                catalogue_df['AGLC SKU'] = weekly_df['SKU'].tolist() if 'SKU' in weekly_df.columns else range(len(catalogue_df))
                catalogue_sku_col = 'AGLC SKU'
                st.warning(f"Created synthetic 'AGLC SKU' column in catalogue_df")
            else:
                catalogue_sku_col = 'AGLC SKU'
        
        if not weekly_sku_col and 'Supplier SKU' in weekly_df.columns:
            weekly_sku_col = 'Supplier SKU'
        elif not weekly_sku_col and 'SKU' in weekly_df.columns:
            weekly_sku_col = 'SKU'
            
        if not stock_qty_col and 'In Stock Qty' in weekly_df.columns:
            stock_qty_col = 'In Stock Qty'
        
        if not catalogue_sku_col or not weekly_sku_col or not stock_qty_col:
            st.error("Still missing critical columns after fallbacks")
            if not catalogue_sku_col:
                catalogue_sku_col = catalogue_df.columns[0]  # Use first column as last resort
                st.warning(f"Using first column '{catalogue_sku_col}' as SKU column")
            if not weekly_sku_col:
                weekly_sku_col = weekly_df.columns[0]  # Use first column as last resort
                st.warning(f"Using first column '{weekly_sku_col}' as Supplier SKU column")
            if not stock_qty_col:
                weekly_df['In Stock Qty'] = 0  # Create a default column
                stock_qty_col = 'In Stock Qty'
                st.warning("Created default 'In Stock Qty' column")
                
        st.warning(f"Falling back to: Catalogue['{catalogue_sku_col}'] = ETL['{weekly_sku_col}'], '{stock_qty_col}'")
    
    # Now perform the merge
    # First, create copies of the columns with standard names for merging
    catalogue_df = catalogue_df.copy()
    weekly_df = weekly_df.copy()
    
    # Rename for merge
    catalogue_df["_merge_key"] = catalogue_df[catalogue_sku_col]
    
    # Process Supplier SKU to extract CNB codes that match AGLC SKU format
    def extract_cnb_code(supplier_sku):
        if pd.isna(supplier_sku):
            return ""
        
        supplier_sku = str(supplier_sku).strip().upper()
        
        # Look for CNB codes in the format CNB-XXXXXX
        if "CNB-" in supplier_sku:
            parts = supplier_sku.split(",")
            for part in parts:
                part = part.strip()
                if part.startswith("CNB-"):
                    return part  # Return just the CNB-XXXXXX part
        
        return supplier_sku
    
    weekly_df["_merge_key"] = weekly_df[weekly_sku_col].apply(extract_cnb_code)
    weekly_df["_stock_qty"] = weekly_df[stock_qty_col]
    
    # Show a sample of the merge keys
    st.info("Sample of merge keys:")
    if len(catalogue_df) > 0:
        st.code(f"Catalogue: {catalogue_df['_merge_key'].head(3).tolist()}")
    if len(weekly_df) > 0:
        st.code(f"ETL output: {weekly_df['_merge_key'].head(3).tolist()}")
    
    # Perform the merge
    merged = (
        catalogue_df
        .merge(
            weekly_df[["_merge_key", "_stock_qty"]],
            left_on="_merge_key",
            right_on="_merge_key",
            how="left"
        )
        .assign(**{"In Stock Qty": lambda d: d["_stock_qty"].fillna(0).astype(int)})
    )
    
    # Drop temporary columns
    merged = merged.drop(columns=["_merge_key", "_stock_qty"], errors="ignore")
    
    # Log merge results
    st.info(f"Merged result has {len(merged)} rows and {len(merged.columns)} columns")
    matched_count = (merged["In Stock Qty"] > 0).sum()
    st.info(f"Found {matched_count} products with stock quantity > 0")

    # 6) Build final in-memory workbook with separate sheets per location
    out_buffer = io.BytesIO()
    
    # Define the Excel date format to use
    excel_date_format = 'yyyy-mm-dd'
    
    # Improve location detection with more alternatives and patterns
    # First, check if we have location data in the weekly_df
    location_col = None
    location_alternatives = [
        "Location", "Store", "Outlet", "Branch", "Site", "Store Name", 
        "Location Name", "Retail Location", "Store Location"
    ]
    
    for col in location_alternatives:
        if col in weekly_df.columns:
            location_col = col
            break
    
    # If not found, try looking for columns with location-related keywords
    if not location_col:
        location_keywords = ["loc", "store", "branch", "site", "outlet"]
        for keyword in location_keywords:
            matching_cols = [col for col in weekly_df.columns if keyword in col.lower()]
            if matching_cols:
                location_col = matching_cols[0]
                break
    
    # Check if the sheet names themselves might represent locations
    # This handles the case where generate_order.py creates one sheet per location
    sheet_locations = [sheet for sheet in available_sheets 
                      if sheet.lower() not in ['all_locations', 'summary', 'info', 'catalogue']]
    
    if location_col and location_col in weekly_df.columns:
        # Use the explicit location column
        locations = weekly_df[location_col].dropna().unique().tolist()
        location_source = f"from column '{location_col}'"
    elif sheet_locations:
        # Use sheet names as locations
        locations = sheet_locations
        location_source = "from sheet names"
        
        # Create a new location column
        weekly_df['Location'] = None
        location_col = 'Location'
        
        # Assign location based on which sheet the data came from
        for loc in locations:
            if loc in all_dfs:
                # Find rows from this sheet in the combined dataframe
                # This is approximate based on the presence of all rows, might need refinement
                # to handle exact matching if sheets have overlapping data
                weekly_df.loc[weekly_df.index[:len(all_dfs[loc])], 'Location'] = loc
    else:
        # No locations found
        locations = []
        location_source = "not found"
    
    # Report findings
    if locations:
        st.success(f"Found {len(locations)} locations {location_source}: {locations}")
    else:
        st.warning("No location information could be detected in the data.")
    
    with pd.ExcelWriter(out_buffer, engine="openpyxl", datetime_format=excel_date_format) as writer:
        if locations:
            # Process each location separately
            for location in locations:
                # Filter weekly_df for just this location
                if location_col in weekly_df.columns:
                    location_df = weekly_df[weekly_df[location_col] == location].copy()
                    if len(location_df) == 0:
                        # Try case-insensitive matching if no rows found
                        if isinstance(location, str) and any(isinstance(val, str) for val in weekly_df[location_col].dropna()):
                            location_df = weekly_df[weekly_df[location_col].str.lower() == location.lower()].copy()
                    
                    st.info(f"Location '{location}' has {len(location_df)} inventory records")
                else:
                    # If we're using sheet names as locations, just use the original sheet data
                    if location in all_dfs:
                        location_df = all_dfs[location].copy()
                        st.info(f"Using sheet '{location}' directly with {len(location_df)} rows")
                    else:
                        # Fallback to empty dataframe
                        location_df = pd.DataFrame(columns=weekly_df.columns)
                        st.warning(f"No data found for location '{location}'")
                
                # Only process if we have data
                if len(location_df) > 0:
                    # Look for the SKU column in this location's data
                    loc_sku_col = weekly_sku_col if weekly_sku_col in location_df.columns else None
                    if not loc_sku_col:
                        # Try to find any SKU-like column
                        for col in supplier_sku_alternatives:
                            if col in location_df.columns:
                                loc_sku_col = col
                                break
                    
                    if not loc_sku_col:
                        sku_cols = [col for col in location_df.columns if "sku" in col.lower()]
                        if sku_cols:
                            loc_sku_col = sku_cols[0]
                    
                    # Look for stock quantity column
                    loc_stock_col = stock_qty_col if stock_qty_col in location_df.columns else None
                    if not loc_stock_col:
                        for col in ["In Stock Qty", "Stock Qty", "Quantity"]:
                            if col in location_df.columns:
                                loc_stock_col = col
                                break
                    
                    if not loc_stock_col:
                        stock_cols = [col for col in location_df.columns if "stock" in col.lower() or "qty" in col.lower()]
                        if stock_cols:
                            loc_stock_col = stock_cols[0]
                    
                    # Only proceed if we found the required columns
                    if loc_sku_col and loc_stock_col:
                        # Create merge keys for this location's data
                        location_df["_merge_key"] = location_df[loc_sku_col].apply(extract_cnb_code)
                        location_df["_stock_qty"] = location_df[loc_stock_col]
                        
                        # Simply extract all columns from the location data
                        etl_columns_to_extract = ["_merge_key", "_stock_qty"] + [
                            col for col in location_df.columns 
                            if col != loc_sku_col and col != loc_stock_col and col != "_merge_key" and col != "_stock_qty"
                        ]
                        
                        # Log which columns we're extracting from the ETL data
                        st.info(f"Extracting all {len(etl_columns_to_extract)} columns from ETL data")
                        
                        # Perform merge for this location - get all columns from both sources
                        etl_extract_df = location_df[etl_columns_to_extract].copy()
                        
                        # Debug the merge operation - check what's in the catalogue dataframe
                        st.info(f"Catalogue columns before merge: {catalogue_df.columns.tolist()}")
                        
                        # Make sure we're not losing any important columns from the order form
                        # Especially look for EachesPerCase
                        if 'EachesPerCase' in catalogue_df.columns:
                            st.success("✅ Found EachesPerCase column in the order form")
                        else:
                            order_form_case_cols = [col for col in catalogue_df.columns if 'case' in col.lower()]
                            if order_form_case_cols:
                                st.info(f"Order form case-related columns: {order_form_case_cols}")
                            else:
                                st.warning("No EachesPerCase or similar column found in order form")
                        
                        # Perform merge but ensure we keep all columns
                        location_merged = (
                            catalogue_df
                            .merge(
                                etl_extract_df,
                                left_on="_merge_key",
                                right_on="_merge_key",
                                how="left",
                                suffixes=('', '_etl')  # Avoid renaming catalogue columns
                            )
                            .assign(**{"In Stock Qty": lambda d: d["_stock_qty"].fillna(0).astype(int)})
                        )
                        
                        # Check if the merge kept all important columns
                        if 'EachesPerCase' in location_merged.columns:
                            st.success("✅ EachesPerCase column preserved in merge result")
                        else:
                            st.warning("EachesPerCase column not found after merge")
                        
                        # Drop temporary merge key and stock qty columns
                        location_merged = location_merged.drop(columns=["_merge_key", "_stock_qty"], errors="ignore")
                        
                        # Check if all catalogue columns are preserved
                        catalogue_cols = set(catalogue_df.columns)
                        merged_cols = set(location_merged.columns)
                        missing_cols = catalogue_cols - merged_cols
                        
                        if missing_cols:
                            st.warning(f"Some columns from the original order form are missing: {missing_cols}")
                            # Try to recover missing columns
                            for col in missing_cols:
                                location_merged[col] = catalogue_df[col]
                            st.info("Recovered missing columns from order form")
                        
                        # Special handling for case sizes - add a fallback method
                        if 'EachesPerCase' not in location_merged.columns:
                            # Look at all columns case-insensitively
                            all_cols_lower = {col.lower(): col for col in location_merged.columns}
                            
                            for possible_name in ['eachespercase', 'eaches per case', 'case size', 'units per case']:
                                if possible_name in all_cols_lower:
                                    actual_col = all_cols_lower[possible_name]
                                    st.success(f"Found case size column with case-insensitive match: '{actual_col}'")
                                    # Create an alias to EachesPerCase
                                    location_merged['EachesPerCase'] = location_merged[actual_col]
                                    break
                                    
                            # Final fallback - if we see columns with both "Case" and a number, use that
                            for col in location_merged.columns:
                                col_lower = col.lower()
                                if 'case' in col_lower and any(str(num) in col for num in range(10)):
                                    st.info(f"Using column '{col}' as potential case size indicator")
                                    try:
                                        # See if it contains numeric values
                                        location_merged[col] = pd.to_numeric(location_merged[col], errors='coerce')
                                        if location_merged[col].dropna().any():
                                            location_merged['EachesPerCase'] = location_merged[col]
                                            st.success(f"Created EachesPerCase column from '{col}'")
                                            break
                                    except:
                                        pass
                            
                            # If we still haven't found it, create a default EachesPerCase based on product type
                            if 'EachesPerCase' not in location_merged.columns:
                                st.warning("Could not find EachesPerCase column after all attempts - creating intelligent defaults")
                                
                                # Default values by category
                                default_case_sizes = {
                                    'Dried Flower': 6,        # Usually 6 per case
                                    'Pre-Roll': 12,           # Usually 12 per case
                                    'Edible': 12,             # Usually 12 per case
                                    'Concentrate': 12,        # Usually 12 per case
                                    'Vaporizer': 10,          # Usually 10 per case
                                    'Beverage': 12,           # Usually 12 per case
                                    'Topical': 12,            # Usually 12 per case
                                    'Accessory': 6,           # Usually 6 per case
                                    'Seeds': 10,              # Usually 10 per case
                                    'Oil': 12,                # Usually 12 per case
                                    'Spray': 12,              # Usually 12 per case
                                    'Capsule': 12,            # Usually 12 per case
                                }
                                
                                # Create the column with category-based defaults
                                if 'Classification' in location_merged.columns:
                                    # Use classification to determine case size
                                    location_merged['EachesPerCase'] = location_merged['Classification'].apply(
                                        lambda x: next((size for category, size in default_case_sizes.items() 
                                                        if category.lower() in str(x).lower()), 12)
                                    )
                                    st.info("Created EachesPerCase column with values based on product classification")
                                else:
                                    # Just use a standard default
                                    location_merged['EachesPerCase'] = 12
                                    st.info("Created EachesPerCase column with default value of 12")
                        
                        # Just use all columns from the merged dataframe
                        all_cols = location_merged.columns.tolist()
                        
                        # Ensure "Order Qty" column exists for the user to fill in
                        if "Order Qty" not in all_cols:
                            # Try to position it after In Stock Qty
                            if "In Stock Qty" in all_cols:
                                insert_pos = all_cols.index("In Stock Qty") + 1
                                location_merged.insert(insert_pos, "Order Qty", "")
                                all_cols.insert(insert_pos, "Order Qty")
                            else:
                                location_merged["Order Qty"] = ""
                                all_cols.append("Order Qty")
                        
                        # Create sheet name from location (ensure it's valid for Excel)
                        sheet_name = str(location)[:31].replace(":", "-").replace("/", "-").replace(" ", "_")
                        
                        # Ensure In Stock Qty is numeric if it exists
                        if "In Stock Qty" in location_merged.columns:
                            try:
                                location_merged["In Stock Qty"] = location_merged["In Stock Qty"].fillna(0).astype(int)
                            except:
                                # Just keep as is if conversion fails
                                pass
                        
                        # Enable deeper debugging if EachesPerCase column isn't found
                        if 'EachesPerCase' not in catalogue_df.columns and 'EachesPerCase' not in location_merged.columns:
                            # This is a critical missing column; provide more details to help debug
                            st.error("EachesPerCase column not found! Enabling debug mode.")
                            st.info(f"All columns in order form (catalogue_df): {sorted(catalogue_df.columns.tolist())}")
                            st.info(f"First 5 rows of order form data:")
                            st.write(catalogue_df.head())
                            
                            # Look for similar columns that might contain the case size information
                            case_related = [col for col in catalogue_df.columns if 'case' in col.lower()]
                            if case_related:
                                st.info(f"Potential case-related columns found: {case_related}")
                                for col in case_related:
                                    st.info(f"Sample values for {col}: {catalogue_df[col].head().tolist()}")
                        
                        # Format date columns properly
                        date_columns = [
                            "First Received Date", "Last Received Date", "Last In Stock Date",
                            "First Received", "Last Received", "Last In Stock",
                            "First Receipt Date", "Last Receipt Date"
                        ]
                        
                        for col in location_merged.columns:
                            if col in date_columns or any(date_term in col.lower() for date_term in ["date", "received", "receipt"]):
                                try:
                                    # Convert to datetime if not already
                                    if not pd.api.types.is_datetime64_dtype(location_merged[col]):
                                        location_merged[col] = pd.to_datetime(location_merged[col], errors='coerce')
                                except:
                                    # Just keep as is if conversion fails
                                    pass
                        
                        # Add order calculation columns - make sure they're visible
                        today = pd.to_datetime('today')
                        days_to_receiving = (receiving_date - today.date()).days
                        
                        # Add calculation columns at the beginning to make them more visible
                        # Create a list of existing columns
                        existing_cols = location_merged.columns.tolist()
                        
                        # Add new columns to the beginning
                        location_merged.insert(0, 'Receiving Date', receiving_date)
                        location_merged.insert(1, 'Days Until Receiving', days_to_receiving)
                        location_merged.insert(2, 'Coverage Period', days_to_receiving + 14)  # receiving + 14 days
                        
                        # Find sales per day column
                        sales_per_day_col = None
                        for col in location_merged.columns:
                            if col.lower() in ['sales/day', 'sales per day', 'daily sales', 'sales_per_day']:
                                sales_per_day_col = col
                                break
                        
                        # Find on order column
                        on_order_col = None
                        for col in location_merged.columns:
                            if col.lower() in ['on order', 'on order qty', 'onorder', 'on_order']:
                                on_order_col = col
                                break
                        
                        # Find case size column - exhaustive list of possible names
                        case_size_col = None
                        case_size_alternatives = [
                            'EachesPerCase',         # Primary expected name from AGLC form
                            'Eaches Per Case',       # Possible variation with spaces
                            'eachespercase',         # Lowercase variation
                            'Eaches_Per_Case',       # Variation with underscores
                            'Case Size',             # Alternative naming
                            'CaseSize',              # Alternative without space
                            'Case_Size',             # Alternative with underscore
                            'Units Per Case',        # Different wording
                            'Units/Case',            # Different wording with slash
                            'Case Quantity',         # Another possible name
                            'Case Qty',              # Abbreviated version
                            'Size',                  # Simple version
                            'UPC',                   # Sometimes used for Units Per Case
                            'Unit/Case',             # Singular variation
                            'Package Size',          # Another way to express it
                            'Pack Size',             # Another way to express it
                            'Pkg Size',              # Abbreviated version
                            'Count Per Case'         # Explicit naming
                        ]
                        
                        for name in case_size_alternatives:
                            if name in location_merged.columns:
                                case_size_col = name
                                break
                                
                        # If still not found, try case-insensitive match
                        if case_size_col is None:
                            for col in location_merged.columns:
                                col_lower = col.lower()
                                if any(alt.lower() == col_lower for alt in case_size_alternatives) or 'case' in col_lower and ('size' in col_lower or 'eaches' in col_lower or 'units' in col_lower):
                                    case_size_col = col
                                    break
                        
                        # If we have sales per day, we can calculate order quantities
                        if sales_per_day_col is not None:
                            # Make numeric
                            try:
                                location_merged[sales_per_day_col] = pd.to_numeric(location_merged[sales_per_day_col], errors='coerce').fillna(0)
                            except:
                                pass
                            
                            # Calculate projected need
                            location_merged['Projected Need'] = location_merged[sales_per_day_col] * location_merged['Coverage Period']
                            
                            # Current inventory including on order
                            if "In Stock Qty" in location_merged.columns:
                                try:
                                    location_merged["In Stock Qty"] = pd.to_numeric(location_merged["In Stock Qty"], errors='coerce').fillna(0)
                                except:
                                    pass
                                
                                if on_order_col is not None:
                                    try:
                                        location_merged[on_order_col] = pd.to_numeric(location_merged[on_order_col], errors='coerce').fillna(0)
                                    except:
                                        pass
                                    location_merged['Current Inventory'] = location_merged["In Stock Qty"] + location_merged[on_order_col]
                                else:
                                    location_merged['Current Inventory'] = location_merged["In Stock Qty"]
                                    location_merged['On Order'] = 0  # Add placeholder
                                
                                # Calculate units needed
                                location_merged['Units Needed'] = (location_merged['Projected Need'] - location_merged['Current Inventory']).clip(lower=0)
                                
                                # Calculate cases needed if we have case size
                                if case_size_col is not None:
                                    try:
                                        location_merged[case_size_col] = pd.to_numeric(location_merged[case_size_col], errors='coerce').fillna(1)
                                    except:
                                        pass
                                    
                                    # Calculate cases needed (to 1 decimal place without rounding up)
                                    location_merged['Cases Needed'] = (location_merged['Units Needed'] / location_merged[case_size_col]).round(1)
                                    
                                    # Set Order Qty to Cases Needed
                                    location_merged['Order Qty'] = location_merged['Cases Needed']
                                else:
                                    # No case size column found, but check if EachesPerCase exists directly
                                    if 'EachesPerCase' in location_merged.columns:
                                        st.info("Using EachesPerCase column from order form for calculations")
                                        try:
                                            location_merged['EachesPerCase'] = pd.to_numeric(location_merged['EachesPerCase'], errors='coerce').fillna(1)
                                            location_merged['Cases Needed'] = (location_merged['Units Needed'] / location_merged['EachesPerCase']).round(1)
                                            location_merged['Order Qty'] = location_merged['Cases Needed']
                                        except Exception as e:
                                            st.error(f"Error using EachesPerCase: {str(e)}")
                                            # Fallback to simple calculation
                                            location_merged['Case Size'] = 1  # Add placeholder
                                            location_merged['Cases Needed'] = location_merged['Units Needed'].round(1)
                                            location_merged['Order Qty'] = location_merged['Cases Needed']
                                    else:
                                        # No case size, just use units
                                        location_merged['Case Size'] = 1  # Add placeholder
                                        location_merged['Cases Needed'] = location_merged['Units Needed'].round(1)
                                        location_merged['Order Qty'] = location_merged['Cases Needed']
                        
                        # Write directly with all calculation columns
                        final_location_df = location_merged
                        
                        # Write to sheet
                        final_location_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Report success with stats
                        match_count = 0
                        if "In Stock Qty" in final_location_df.columns:
                            match_count = (pd.to_numeric(final_location_df["In Stock Qty"], errors='coerce') > 0).sum()
                        
                        # Log detailed column information
                        st.success(f"Created sheet for location '{location}' with {len(final_location_df)} products ({match_count} with stock > 0)")
                        st.info(f"Sheet contains {len(final_location_df.columns)} columns")
                    else:
                        st.error(f"Could not find required columns for location '{location}'. " 
                                f"Need SKU column (found: {loc_sku_col}) and stock column (found: {loc_stock_col}).")
            
            # Also create a combined sheet with all data
            combined_sheet = "All Locations"
            
            # Drop any temporary columns
            if "_merge_key" in merged.columns:
                merged = merged.drop(columns=["_merge_key"], errors="ignore")
            if "_stock_qty" in merged.columns:
                merged = merged.drop(columns=["_stock_qty"], errors="ignore")
            
            # Ensure Order Qty column exists
            if "Order Qty" not in merged.columns:
                # Insert after In Stock Qty if it exists
                if "In Stock Qty" in merged.columns:
                    # Find the position of In Stock Qty
                    cols = merged.columns.tolist()
                    pos = cols.index("In Stock Qty") + 1
                    # Insert Order Qty after it
                    merged.insert(pos, "Order Qty", "")
                else:
                    # Add at the end
                    merged["Order Qty"] = ""
            
            # Ensure In Stock Qty is numeric if it exists
            if "In Stock Qty" in merged.columns:
                try:
                    merged["In Stock Qty"] = merged["In Stock Qty"].fillna(0).astype(int)
                except:
                    # Keep as is if conversion fails
                    pass
            
            # Format date columns properly
            date_columns = [
                "First Received Date", "Last Received Date", "Last In Stock Date",
                "First Received", "Last Received", "Last In Stock",
                "First Receipt Date", "Last Receipt Date"
            ]
            
            for col in merged.columns:
                if col in date_columns or any(date_term in col.lower() for date_term in ["date", "received", "receipt"]):
                    try:
                        # Convert to datetime if not already
                        if not pd.api.types.is_datetime64_dtype(merged[col]):
                            merged[col] = pd.to_datetime(merged[col], errors='coerce')
                    except:
                        # Just keep as is if conversion fails
                        pass
            
            # Write the combined data with all columns
            merged.to_excel(writer, sheet_name=combined_sheet, index=False)
            st.info(f"Created combined data in sheet: '{combined_sheet}' with {len(merged.columns)} columns")
            
        else:
            # If no locations found, just use the original merged data
            sheet_name = "Catalogue"  # Use the British/Canadian spelling with "ue"
            
            # Drop any temporary columns
            if "_merge_key" in merged.columns:
                merged = merged.drop(columns=["_merge_key"], errors="ignore")
            if "_stock_qty" in merged.columns:
                merged = merged.drop(columns=["_stock_qty"], errors="ignore")
            
            # Ensure Order Qty column exists
            if "Order Qty" not in merged.columns:
                # Insert after In Stock Qty if it exists
                if "In Stock Qty" in merged.columns:
                    # Find the position of In Stock Qty
                    cols = merged.columns.tolist()
                    pos = cols.index("In Stock Qty") + 1
                    # Insert Order Qty after it
                    merged.insert(pos, "Order Qty", "")
                else:
                    # Add at the end
                    merged["Order Qty"] = ""
            
            # Ensure In Stock Qty is numeric if it exists
            if "In Stock Qty" in merged.columns:
                try:
                    merged["In Stock Qty"] = merged["In Stock Qty"].fillna(0).astype(int)
                except:
                    # Keep as is if conversion fails
                    pass
            
            # Format date columns properly
            date_columns = [
                "First Received Date", "Last Received Date", "Last In Stock Date",
                "First Received", "Last Received", "Last In Stock",
                "First Receipt Date", "Last Receipt Date"
            ]
            
            for col in merged.columns:
                if col in date_columns or any(date_term in col.lower() for date_term in ["date", "received", "receipt"]):
                    try:
                        # Convert to datetime if not already
                        if not pd.api.types.is_datetime64_dtype(merged[col]):
                            merged[col] = pd.to_datetime(merged[col], errors='coerce')
                    except:
                        # Just keep as is if conversion fails
                        pass
            
            # Write the data with all columns
            merged.to_excel(writer, sheet_name=sheet_name, index=False)
            st.info(f"Created data in sheet: '{sheet_name}' with {len(merged.columns)} columns (no location data found)")
        
        # Add a debug info sheet
        debug_info = pd.DataFrame({
            "Name": [
                "Generation Date", 
                "Historical Days", 
                "Excluded Today",
                "Locations Found",
                "Total Products",
                "Catalogue Source",
                "Data Source"
            ],
            "Value": [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                hist_days,
                "Yes" if exclude_today else "No",
                ", ".join(str(loc) for loc in locations) if locations else "None",
                len(catalogue_df),
                "Manual Upload" if 'upload_file' in locals() and upload_file else "Automatic Download",
                "Combined ETL sheets"
            ]
        })
        debug_info.to_excel(writer, sheet_name="Info", index=False)
    out_buffer.seek(0)

    # 7) Offer a single download
    today_stamp = datetime.now().strftime("%Y%m%d")
    
    # Create a more descriptive filename based on the locations
    if locations:
        if len(locations) == 1:
            # Single location
            location_str = str(locations[0]).replace(" ", "_")[:20]
            file_name = f"OrderForm_{location_str}_{today_stamp}.xlsx"
        else:
            # Multiple locations
            file_name = f"OrderForm_{len(locations)}_Locations_{today_stamp}.xlsx"
    else:
        # No locations found
        file_name = f"Compiled_OrderForm_{today_stamp}.xlsx"
    
    st.download_button(
        f"⬇️ Download Order Form ({len(locations) if locations else 'All'} locations)",
        data=out_buffer,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
    # Add instructions for using the downloaded file
    st.success("✅ Order form generation complete!")
    
    st.markdown(f"""
    ### Next Steps:
    
    1. Open the downloaded Excel file
    2. For each location, go to the corresponding sheet
    3. Review the automated order calculations:
       - **Receiving Date**: {receiving_date.strftime('%Y-%m-%d')} (selected by you)
       - **Coverage Period**: {(receiving_date - pd.to_datetime('today').date()).days + 14} days (days until receiving + 14 days)
       - **Projected Need**: Sales/day × Coverage Period
       - **Current Inventory**: In Stock Qty + On Order
       - **Units Needed**: Projected Need - Current Inventory
       - **Cases Needed**: Units Needed ÷ Case Size (rounded to 1 decimal place)
    4. The "Order Qty" column has been automatically filled with suggested values
    5. Review and adjust the "Order Qty" values as needed
    6. Save the file and submit according to AGLC instructions
    
    Each sheet contains the complete order form for a single location with order calculations.
    The "All Locations" sheet shows all products across all locations.
    """)
    
    if locations:
        st.info(f"Your file contains sheets for these locations: {', '.join(str(loc) for loc in locations)}")
    else:
        st.warning("No location data was found in the inventory. The order form has been created with a single Catalogue sheet.")
